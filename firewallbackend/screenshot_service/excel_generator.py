"""
Générateur de fichiers Excel avec screenshots intégrés
"""

import os
import base64
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

def generate_excel_with_screenshot(ip_address, protocol, url, screenshot_base64, width, height, user=None):
    """
    Génère un fichier Excel avec le screenshot intégré
    
    Args:
        ip_address: Adresse IP du firewall
        protocol: Protocole utilisé (http/https)
        url: URL complète
        screenshot_base64: Screenshot encodé en base64
        width: Largeur du screenshot
        height: Hauteur du screenshot
        user: Utilisateur qui a généré le rapport
    
    Returns:
        tuple: (file_path, filename) ou (None, None) en cas d'erreur
    """
    try:
        # Créer le dossier pour les rapports Excel
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'screenshot_reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Générer le nom du fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"screenshot_report_{ip_address}_{timestamp}.xlsx"
        filepath = os.path.join(reports_dir, filename)
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Screenshot Report"
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        info_font = Font(bold=True)
        
        # En-tête principal
        ws.merge_cells('A1:D1')
        header_cell = ws['A1']
        header_cell.value = f"Rapport Screenshot - Firewall {ip_address}"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Informations du rapport
        row = 3
        info_data = [
            ("Adresse IP:", ip_address),
            ("Protocole:", protocol.upper()),
            ("URL:", url),
            ("Dimensions:", f"{width} x {height} pixels"),
            ("Date de génération:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Généré par:", user or "Système")
        ]
        
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = info_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Ajouter le screenshot
        if screenshot_base64:
            try:
                # Décoder l'image base64
                image_data = base64.b64decode(screenshot_base64)
                image = Image(io.BytesIO(image_data))
                
                # Redimensionner l'image pour qu'elle rentre bien dans Excel (taille agrandie x2)
                # Garder les proportions
                max_width = 1200
                max_height = 800
                
                if image.width > max_width or image.height > max_height:
                    ratio = min(max_width / image.width, max_height / image.height)
                    image.width = int(image.width * ratio)
                    image.height = int(image.height * ratio)
                
                # Ajouter l'image à la feuille (après les informations)
                image_row = row + 2
                ws.add_image(image, f'A{image_row}')
                
                # Ajuster la hauteur de la ligne pour l'image
                ws.row_dimensions[image_row].height = int(image.height * 0.75)  # Convertir pixels en points
                
                logger.info(f"Screenshot ajouté au fichier Excel: {filename}")
                
            except ImportError as e:
                logger.error(f"Pillow non installé: {e}")
                # Message d'erreur plus informatif
                error_row = row + 2
                ws.cell(row=error_row, column=1, value="⚠️ Erreur: Pillow non installé")
                ws.cell(row=error_row, column=1).font = Font(color="FF0000", bold=True)
                ws.cell(row=error_row + 1, column=1, value="Pour ajouter l'image, installez Pillow: pip install Pillow")
                ws.cell(row=error_row + 1, column=1).font = Font(color="FF6600")
                
            except Exception as e:
                logger.error(f"Erreur lors de l'ajout du screenshot: {e}")
                # Ajouter un message d'erreur
                error_row = row + 2
                ws.cell(row=error_row, column=1, value="❌ Erreur: Impossible d'ajouter le screenshot")
                ws.cell(row=error_row, column=1).font = Font(color="FF0000", bold=True)
                ws.cell(row=error_row + 1, column=1, value=f"Détails: {str(e)}")
                ws.cell(row=error_row + 1, column=1).font = Font(color="FF6600")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Sauvegarder le fichier
        wb.save(filepath)
        
        logger.info(f"Fichier Excel généré avec succès: {filepath}")
        return filepath, filename
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du fichier Excel: {e}")
        return None, None

def cleanup_old_reports(days_old=7):
    """
    Nettoie les anciens rapports Excel
    
    Args:
        days_old: Nombre de jours après lesquels supprimer les fichiers
    """
    try:
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'screenshot_reports')
        if not os.path.exists(reports_dir):
            return
        
        current_time = datetime.now().timestamp()
        cutoff_time = current_time - (days_old * 24 * 60 * 60)
        
        deleted_count = 0
        for filename in os.listdir(reports_dir):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(reports_dir, filename)
                if os.path.getmtime(filepath) < cutoff_time:
                    os.remove(filepath)
                    deleted_count += 1
        
        if deleted_count > 0:
            logger.info(f"Nettoyage: {deleted_count} anciens rapports supprimés")
            
    except Exception as e:
        logger.error(f"Erreur lors du nettoyage des anciens rapports: {e}")
