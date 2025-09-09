from django.shortcuts import render
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError
import asyncio
import json
import logging
import os
import uuid
from datetime import datetime, timedelta
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from auth_service.models import SSHUser
from auth_service.utils.crypto import decrypt_ssh_data

from .models import (
    EmailLog, AutomatedEmailSchedule, AutomatedEmailExecution, 
    CommandExecutionResult, CommandTemplate
)
from .serializers import (
    EmailLogSerializer, AutomatedEmailScheduleSerializer, 
    AutomatedEmailExecutionSerializer, CommandExecutionResultSerializer,
    EmailScheduleListSerializer, CommandTemplateSerializer
)
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from django.db.models import Sum, Count, Q
import asyncio

User = get_user_model()
logger = logging.getLogger(__name__)


class EmailLogViewSet(viewsets.ModelViewSet):
    queryset = EmailLog.objects.all()
    serializer_class = EmailLogSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]


class AdminSendEmailView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def post(self, request):
        subject = request.data.get('subject', '')
        message = request.data.get('message', '')
        recipient_ids = request.data.get('recipient_ids', [])
        
        if not subject or not message or not recipient_ids:
            return Response({
                'error': 'Subject, message, and recipient_ids are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        users = User.objects.filter(id__in=recipient_ids)
        recipient_list = [user.email for user in users if user.email]
        
        if not recipient_list:
            return Response({
                'error': 'No valid email addresses found'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=None,  # Uses DEFAULT_FROM_EMAIL
                recipient_list=recipient_list,
                fail_silently=False,
            )
            
            # Log the emails sent
            for user in users:
                if user.email:
                    EmailLog.objects.create(
                        recipient=user.email,
                        subject=subject,
                        content=message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        smtp_host=settings.EMAIL_HOST,
                        smtp_port=settings.EMAIL_PORT,
                        status='sent'
                    )

            return Response({'success': True, 'recipients': recipient_list})
        except Exception as e:
            logger.error(f"Email sending failed: {str(e)}")
            return Response({
                'error': f'Failed to send email: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AutomatedEmailScheduleViewSet(viewsets.ModelViewSet):
    queryset = AutomatedEmailSchedule.objects.all()
    serializer_class = AutomatedEmailScheduleSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return AutomatedEmailSchedule.objects.filter(created_by=self.request.user)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_serializer_class(self):
        if self.action == 'list':
            return EmailScheduleListSerializer
        return AutomatedEmailScheduleSerializer

    @action(detail=True, methods=['post'])
    def execute_now(self, request, pk=None):
        """Exécuter immédiatement un planning d'email"""
        schedule = self.get_object()
        
        try:
            execution = AutomatedEmailExecution.objects.create(
                schedule=schedule,
                status='pending'
            )
            
            # Lancer l'exécution en arrière-plan avec threading
            import threading
            thread = threading.Thread(target=self.execute_email_schedule, args=(execution.id,))
            thread.daemon = True
            thread.start()
            
            return Response({
                'message': 'Exécution démarrée',
                'execution_id': execution.id
            })
        except Exception as e:
            return Response({
                'error': f"Erreur lors du démarrage de l'exécution: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Activer/désactiver un planning"""
        schedule = self.get_object()
        schedule.is_active = not schedule.is_active
        schedule.save()
        
        return Response({
            'message': f'Planning {"activé" if schedule.is_active else "désactivé"}',
            'is_active': schedule.is_active
        })

    def execute_email_schedule(self, execution_id):
        """Exécuter un planning d'email avec exécution de commandes"""
        try:
            logger.info(f"🚀 [EMAIL_SCHEDULE] Début d'exécution du planning {execution_id}")
            
            execution = AutomatedEmailExecution.objects.get(id=execution_id)
            execution.status = 'running'
            execution.started_at = timezone.now()
            execution.save()
            
            schedule = execution.schedule
            logger.info(f"📋 [EMAIL_SCHEDULE] Planning: {schedule.name} (ID: {schedule.id})")
            
            # Récupérer l'utilisateur administrateur qui a créé le planning
            admin_user = schedule.created_by
            logger.info(f"👤 [EMAIL_SCHEDULE] Admin user: {admin_user.username} (ID: {admin_user.id})")
            
            # Récupérer les destinataires
            recipients = schedule.get_recipients()
            logger.info(f"📧 [EMAIL_SCHEDULE] Destinataires trouvés: {len(recipients)} utilisateurs")
            for recipient in recipients:
                logger.info(f"   - {recipient.username} ({recipient.email})")
            
            # Exécuter les commandes sur les firewalls avec les identifiants de l'admin
            logger.info(f"🔧 [EMAIL_SCHEDULE] Début d'exécution des commandes sur les firewalls...")
            command_results = self.execute_firewall_commands(schedule, execution, admin_user)
            
            # Préparer le contenu de l'email avec les résultats
            logger.info(f"📝 [EMAIL_SCHEDULE] Préparation du contenu de l'email...")
            email_content = self.prepare_email_content(schedule, command_results)
            
            # Générer le fichier Excel avec les résultats
            logger.info(f"📊 [EMAIL_SCHEDULE] Génération du fichier Excel...")
            excel_filepath, excel_filename = self.generate_excel_report(schedule, command_results)
            logger.info(f"📁 [EMAIL_SCHEDULE] Fichier Excel généré: {excel_filename}")
            
            # Envoyer les emails
            logger.info(f"📤 [EMAIL_SCHEDULE] Début d'envoi des emails...")
            emails_sent = 0
            emails_failed = 0
            
            for recipient in recipients:
                if recipient.email:
                    logger.info(f"📧 [EMAIL_SCHEDULE] Envoi d'email à {recipient.username} ({recipient.email})")
                    try:
                        # Personnaliser le contenu pour chaque destinataire
                        personalized_content = email_content.replace(
                            '{{USER_NAME}}', recipient.get_full_name() or recipient.username
                        )
                        logger.info(f"   📝 Contenu personnalisé pour {recipient.username}")
                        
                        # Envoyer l'email avec pièce jointe Excel
                        logger.info(f"   📤 Tentative d'envoi d'email...")
                        self.send_email_with_attachment(
                            subject=schedule.email_subject,
                            message=personalized_content,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[recipient.email],
                            attachment_path=excel_filepath,
                            attachment_filename=excel_filename
                        )
                        
                        # Log de l'email
                        EmailLog.objects.create(
                            recipient=recipient.email,
                            subject=schedule.email_subject,
                            content=personalized_content,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            smtp_host=settings.EMAIL_HOST,
                            smtp_port=settings.EMAIL_PORT,
                            status='sent'
                        )
                        
                        emails_sent += 1
                        logger.info(f"   ✅ Email envoyé avec succès à {recipient.email}")
                        
                    except Exception as e:
                        logger.error(f"❌ [EMAIL_SCHEDULE] Échec d'envoi d'email à {recipient.email}: {str(e)}")
                        emails_failed += 1
                        
                        EmailLog.objects.create(
                            recipient=recipient.email,
                            subject=schedule.email_subject,
                            content=email_content,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            smtp_host=settings.EMAIL_HOST,
                            smtp_port=settings.EMAIL_PORT,
                            status='failed',
                            error_message=str(e)
                        )
                        logger.info(f"   ❌ Email échoué pour {recipient.email}")
            
            # Mettre à jour l'exécution
            logger.info(f"📊 [EMAIL_SCHEDULE] Résumé de l'exécution:")
            logger.info(f"   - Emails envoyés: {emails_sent}")
            logger.info(f"   - Emails échoués: {emails_failed}")
            logger.info(f"   - Commandes réussies: {command_results['successful']}")
            logger.info(f"   - Commandes échouées: {command_results['failed']}")
            
            execution.status = 'completed'
            execution.completed_at = timezone.now()
            execution.emails_sent = emails_sent
            execution.emails_failed = emails_failed
            execution.commands_executed = command_results['successful']
            execution.commands_failed = command_results['failed']
            execution.duration_seconds = (execution.completed_at - execution.started_at).total_seconds()
            execution.save()
            
            # Mettre à jour le planning
            schedule.last_sent = timezone.now()
            schedule.next_send = schedule.calculate_next_send()
            schedule.save()
            
            logger.info(f"✅ [EMAIL_SCHEDULE] Exécution terminée avec succès!")
            
        except Exception as e:
            logger.error(f"Error executing email schedule {execution_id}: {str(e)}")
            try:
                execution = AutomatedEmailExecution.objects.get(id=execution_id)
                execution.status = 'failed'
                execution.error_message = str(e)
                execution.completed_at = timezone.now()
                execution.save()
            except:
                pass

    def execute_firewall_commands(self, schedule, execution, admin_user=None):
        """Exécuter les commandes sur les firewalls via WebSocket"""
        logger.info(f"🔧 [FIREWALL_COMMANDS] Début d'exécution des commandes")
        logger.info(f"   - Planning: {schedule.name}")
        logger.info(f"   - Admin user: {admin_user.username if admin_user else 'None'}")
        logger.info(f"   - Firewalls: {schedule.firewalls.count()}")
        logger.info(f"   - Commandes: {len(schedule.commands_to_execute)}")
        
        # Vérifier s'il y a des firewalls associés
        firewalls_list = list(schedule.firewalls.all())
        if not firewalls_list:
            logger.warning(f"⚠️ [FIREWALL_COMMANDS] AUCUN FIREWALL ASSOCIÉ AU PLANNING!")
            logger.warning(f"   - Planning ID: {schedule.id}")
            logger.warning(f"   - Nom du planning: {schedule.name}")
            logger.warning(f"   - Firewalls sélectionnés: {schedule.selected_firewalls}")
            logger.warning(f"   - Types de firewalls: {schedule.firewall_types}")
            
            # Retourner un résultat vide mais valide
            return {
                'successful': 0,
                'failed': 0,
                'results': []
            }
        
        logger.info(f"   ✅ [FIREWALL_COMMANDS] Firewalls trouvés: {len(firewalls_list)}")
        for fw in firewalls_list:
            logger.info(f"      - {fw.name} ({fw.ip_address}) - Type: {fw.firewall_type.name}")
        
        channel_layer = get_channel_layer()
        successful_commands = 0
        failed_commands = 0
        command_results = []  # Pour collecter les résultats détaillés
        
        for firewall in firewalls_list:
            logger.info(f"🖥️ [FIREWALL_COMMANDS] Traitement du firewall: {firewall.name} ({firewall.ip_address})")
            for command_data in schedule.commands_to_execute:
                command = command_data.get('command', '')
                command_type = command_data.get('type', 'general')
                
                logger.info(f"   💻 [FIREWALL_COMMANDS] Exécution de la commande: {command} (type: {command_type})")
                
                # Créer l'enregistrement de résultat
                result = CommandExecutionResult.objects.create(
                    execution=execution,
                    firewall=firewall,
                    command=command,
                    command_type=command_type,
                    status='executing',
                    started_at=timezone.now()
                )
                logger.info(f"   📝 [FIREWALL_COMMANDS] Enregistrement de résultat créé: {result.id}")
                
                try:
                    # Envoyer la commande via WebSocket avec les identifiants de l'admin
                    self.send_command_via_websocket(
                        channel_layer, firewall, command, result, admin_user
                    )
                    successful_commands += 1
                    
                    # Collecter le résultat pour l'Excel
                    command_results.append({
                        'firewall_name': firewall.name,
                        'firewall_ip': firewall.ip_address,
                        'firewall_type': firewall.firewall_type.name,
                        'command': command,
                        'command_type': command_type,
                        'status': 'success',
                        'output': result.output or f"Commande exécutée avec succès sur {firewall.name}",
                        'execution_time': result.completed_at or timezone.now(),
                        'duration': (result.completed_at - result.started_at).total_seconds() if result.completed_at else 0
                    })
                    
                except Exception as e:
                    logger.error(f"Command execution failed for firewall {firewall.name}: {str(e)}")
                    result.status = 'failed'
                    result.error_output = str(e)
                    result.completed_at = timezone.now()
                    result.save()
                    failed_commands += 1
                    
                    # Collecter le résultat d'erreur pour l'Excel
                    command_results.append({
                        'firewall_name': firewall.name,
                        'firewall_ip': firewall.ip_address,
                        'firewall_type': firewall.firewall_type.name,
                        'command': command,
                        'command_type': command_type,
                        'status': 'failed',
                        'output': str(e),
                        'execution_time': result.completed_at or timezone.now(),
                        'duration': (result.completed_at - result.started_at).total_seconds() if result.completed_at else 0
                    })
        
        return {
            'successful': successful_commands,
            'failed': failed_commands,
            'results': command_results
        }

    def send_command_via_websocket(self, channel_layer, firewall, command, result, admin_user=None):
        """Envoyer une commande via WebSocket et attendre le résultat"""
        logger.info(f"🔌 [WEBSOCKET] Début d'envoi de commande via WebSocket")
        logger.info(f"   - Firewall: {firewall.name} ({firewall.ip_address})")
        logger.info(f"   - Commande: {command}")
        logger.info(f"   - Admin user: {admin_user.username if admin_user else 'None'}")
        
        try:
            # Utiliser le service WebSocket existant pour exécuter la commande
            logger.info(f"   🔄 [WEBSOCKET] Appel du service WebSocket...")
            firewall_response = self.execute_command_via_ssh_session(firewall, command, admin_user)
            
            logger.info(f"   📥 [WEBSOCKET] Réponse reçue: {firewall_response[:100]}...")
            
            # Mettre à jour le résultat avec la vraie réponse du firewall
            result.status = 'completed'
            result.output = firewall_response
            result.completed_at = timezone.now()
            result.save()
            
            logger.info(f"   ✅ [WEBSOCKET] Résultat sauvegardé avec succès")
            
        except Exception as e:
            logger.error(f"❌ [WEBSOCKET] Erreur lors de l'exécution de la commande sur {firewall.name}: {str(e)}")
            result.status = 'failed'
            result.error_output = str(e)
            result.completed_at = timezone.now()
            result.save()
            logger.info(f"   ❌ [WEBSOCKET] Résultat d'erreur sauvegardé")
            raise e

    def execute_command_via_ssh_session(self, firewall, command, admin_user=None):
        """
        Exécuter une commande via SSH avec session interactive
        """
        import asyncio
        from websocket_service.models import TerminalSession, TerminalCommand
        
        try:
            # Créer une session terminal
            session_id = f"email_exec_{uuid.uuid4().hex[:8]}"
            session = TerminalSession.objects.create(
                user=admin_user,
                firewall=firewall,
                session_id=session_id,
                is_active=True
            )
            
            # Créer une commande terminal
            command_id = f"cmd_{uuid.uuid4().hex[:8]}"
            terminal_command = TerminalCommand.objects.create(
                session=session,
                command=command,
                command_id=command_id,
                status='executing'
            )
            
            # Utiliser le service SSH interactif
            from websocket_service.ssh_session_manager import execute_command_via_ssh
            
            # Exécuter la commande
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            try:
                output = loop.run_until_complete(
                    execute_command_via_ssh(firewall, command, command_id, admin_user)
                )
                return output
            finally:
                loop.close()
                
        except Exception as e:
            logger.error(f"Erreur d'exécution: {str(e)}")
            return f"Erreur d'exécution: {str(e)}"
        finally:
            # Nettoyer la session
            try:
                if 'session' in locals():
                    session.is_active = False
                    session.save()
            except:
                pass

    def get_command_output_from_session(self, session, terminal_command):
        """Récupérer la sortie d'une commande depuis la session terminal"""
        try:
            # Cette méthode devrait récupérer la sortie réelle de la commande
            # Pour l'instant, on simule une sortie basée sur le type de firewall
            firewall_type = session.firewall.firewall_type.name.lower()
            
            if 'cisco' in firewall_type:
                return f"Router# {terminal_command.command}\nCisco IOS Software, Version 15.2(4)S7\nCisco ASR1000 Series Routers\nProcessor: 2.5 GHz"
            elif 'fortinet' in firewall_type:
                return f"FortiGate # {terminal_command.command}\nFortiGate-60F v6.4.5\nSerial Number: FG60FTK1234567890\nStatus: online"
            elif 'palo' in firewall_type:
                return f"PA-VM # {terminal_command.command}\nPalo Alto Networks PAN-OS v10.1.0\nModel: PA-VM\nSerial: 0123456789012345"
            else:
                return f"Command executed: {terminal_command.command}\nOutput: Command completed successfully"
                
        except Exception as e:
            logger.error(f"Error getting command output: {str(e)}")
            return f"Erreur lors de la récupération de la sortie: {str(e)}"

    # MÉTHODE SSH DIRECTE - COMMENTÉE POUR UTILISER UNIQUEMENT WEBSOCKET
    # def execute_ssh_command(self, firewall, command, admin_user=None):
    #     """Exécuter une commande SSH sur le firewall et récupérer la réponse (MÉTHODE COMMENTÉE)"""
    #     import paramiko
    #     import time
    #     from auth_service.models import SSHUser
    #     from auth_service.utils.crypto import decrypt_ssh_data
    #     
    #     try:
    #         # Récupérer les informations SSH de l'utilisateur administrateur
    #         ssh_user = None
    #         ssh_password = None
    #         ssh_port = getattr(firewall, 'ssh_port', 22)
    #         
    #         if admin_user:
    #             try:
    #                 ssh_credentials = SSHUser.objects.get(user=admin_user)
    #                 ssh_user = ssh_credentials.ssh_username
    #                 ssh_password = decrypt_ssh_data(ssh_credentials.ssh_password)
    #                 logger.info(f"Using admin user SSH credentials for firewall {firewall.name}")
    #             except SSHUser.DoesNotExist:
    #                 logger.warning(f"No SSH credentials found for admin user {admin_user.username}")
    #                 return "Erreur: Aucun identifiant SSH configuré pour l'utilisateur administrateur"
    #             except Exception as e:
    #                 logger.error(f"Error decrypting SSH credentials: {str(e)}")
    #                 return f"Erreur de déchiffrement des identifiants SSH: {str(e)}"
    #         else:
    #             return "Erreur: Utilisateur administrateur non spécifié"
    #         
    #         # Créer une connexion SSH
    #         ssh_client = paramiko.SSHClient()
    #         ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    #         
    #         # Se connecter au firewall
    #         ssh_client.connect(
    #             hostname=firewall.ip_address,
    #             port=ssh_port,
    #             username=ssh_user,
    #             password=ssh_password,
    #             timeout=10
    #         )
    #         
    #         # Exécuter la commande
    #         stdin, stdout, stderr = ssh_client.exec_command(command, timeout=30)
    #         
    #         # Récupérer la sortie
    #         output = stdout.read().decode('utf-8', errors='ignore')
    #         error_output = stderr.read().decode('utf-8', errors='ignore')
    #         
    #         # Fermer la connexion
    #         ssh_client.close()
    #         
    #         # Retourner la réponse
    #         if error_output:
    #             return f"Erreur: {error_output}\nSortie: {output}"
    #         else:
    #             return output.strip() if output else "Commande exécutée sans sortie"
    #             
    #     except paramiko.AuthenticationException:
    #         return "Erreur d'authentification SSH - Vérifiez les identifiants de l'utilisateur administrateur"
    #     except paramiko.SSHException as e:
    #         return f"Erreur SSH: {str(e)}"
    #     except Exception as e:
    #         return f"Erreur de connexion: {str(e)}"

    def generate_excel_report(self, schedule, command_results):
        """Générer un fichier Excel avec les résultats des commandes"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import os
        from django.conf import settings
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        
        # Page 1: Résumé des commandes
        ws1 = wb.active
        ws1.title = "Résumé des Commandes"
        
        # Page 2: Réponses détaillées des firewalls
        ws2 = wb.create_sheet("Réponses Firewalls")
        
        # Définir les styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # === PAGE 1: RÉSUMÉ ===
        headers1 = [
            'Firewall', 'Adresse IP', 'Type', 'Commande', 'Type Commande', 
            'Statut', 'Résultat', 'Date d\'exécution', 'Durée (s)'
        ]
        
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Données de la page 1
        for row, result in enumerate(command_results['results'], 2):
            ws1.cell(row=row, column=1, value=result['firewall_name'])
            ws1.cell(row=row, column=2, value=result['firewall_ip'])
            ws1.cell(row=row, column=3, value=result['firewall_type'])
            ws1.cell(row=row, column=4, value=result['command'])
            ws1.cell(row=row, column=5, value=result['command_type'])
            
            # Statut avec couleur
            status_cell = ws1.cell(row=row, column=6, value=result['status'])
            if result['status'] == 'success':
                status_cell.fill = success_fill
            else:
                status_cell.fill = error_fill
            
            # Résumé du résultat (premiers caractères)
            output_summary = result['output'][:100] + "..." if len(result['output']) > 100 else result['output']
            ws1.cell(row=row, column=7, value=output_summary)
            ws1.cell(row=row, column=8, value=result['execution_time'].strftime('%Y-%m-%d %H:%M:%S'))
            ws1.cell(row=row, column=9, value=result['duration'])
        
        # === PAGE 2: RÉPONSES DÉTAILLÉES ===
        headers2 = [
            'Firewall', 'Adresse IP', 'Commande', 'Statut', 'Réponse complète du firewall', 
            'Date d\'exécution', 'Durée (s)'
        ]
        
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Données de la page 2 avec réponses complètes
        for row, result in enumerate(command_results['results'], 2):
            ws2.cell(row=row, column=1, value=result['firewall_name'])
            ws2.cell(row=row, column=2, value=result['firewall_ip'])
            ws2.cell(row=row, column=3, value=result['command'])
            
            # Statut avec couleur
            status_cell = ws2.cell(row=row, column=4, value=result['status'])
            if result['status'] == 'success':
                status_cell.fill = success_fill
            else:
                status_cell.fill = error_fill
            
            # Réponse complète du firewall
            ws2.cell(row=row, column=5, value=result['output'])
            ws2.cell(row=row, column=6, value=result['execution_time'].strftime('%Y-%m-%d %H:%M:%S'))
            ws2.cell(row=row, column=7, value=result['duration'])
        
        # Ajuster la largeur des colonnes pour les deux pages
        for ws in [ws1, ws2]:
            for col in range(1, max(len(headers1), len(headers2)) + 1):
                if col == 5:  # Colonne des réponses complètes
                    ws.column_dimensions[get_column_letter(col)].width = 50
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Créer le dossier media s'il n'existe pas
        media_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(media_dir, exist_ok=True)
        
        # Nom du fichier
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"firewall_commands_report_{schedule.name}_{timestamp}.xlsx"
        filepath = os.path.join(media_dir, filename)
        
        # Sauvegarder le fichier
        wb.save(filepath)
        
        return filepath, filename

    def send_email_with_attachment(self, subject, message, from_email, recipient_list, attachment_path, attachment_filename):
        """Envoyer un email avec une pièce jointe"""
        from django.core.mail import EmailMessage
        import os
        
        logger.info(f"📧 [EMAIL_ATTACHMENT] Début d'envoi d'email avec pièce jointe")
        logger.info(f"   - Sujet: {subject}")
        logger.info(f"   - De: {from_email}")
        logger.info(f"   - À: {recipient_list}")
        logger.info(f"   - Pièce jointe: {attachment_filename}")
        logger.info(f"   - Chemin fichier: {attachment_path}")
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=from_email,
            to=recipient_list,
        )
        
        # Attacher le fichier Excel
        if os.path.exists(attachment_path):
            logger.info(f"   📎 [EMAIL_ATTACHMENT] Fichier trouvé, ajout de la pièce jointe...")
            with open(attachment_path, 'rb') as f:
                file_content = f.read()
                email.attach(attachment_filename, file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            logger.info(f"   ✅ [EMAIL_ATTACHMENT] Pièce jointe ajoutée ({len(file_content)} bytes)")
        else:
            logger.warning(f"   ⚠️ [EMAIL_ATTACHMENT] Fichier non trouvé: {attachment_path}")
        
        logger.info(f"   📤 [EMAIL_ATTACHMENT] Envoi de l'email...")
        email.send()
        logger.info(f"   ✅ [EMAIL_ATTACHMENT] Email envoyé avec succès")

    def prepare_email_content(self, schedule, command_results):
        """Préparer le contenu de l'email avec les résultats des commandes"""
        template = schedule.email_template
        
        # Remplacer les variables dans le template
        content = template.replace('{{TOTAL_COMMANDS}}', str(command_results['successful'] + command_results['failed']))
        content = content.replace('{{SUCCESSFUL_COMMANDS}}', str(command_results['successful']))
        content = content.replace('{{FAILED_COMMANDS}}', str(command_results['failed']))
        content = content.replace('{{EXECUTION_DATE}}', timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Ajouter une note sur le fichier Excel attaché
        excel_note = "\n\n📎 Un fichier Excel avec deux pages est attaché à cet email :"
        excel_note += "\n   • Page 1 : Résumé des commandes"
        excel_note += "\n   • Page 2 : Réponses complètes des firewalls"
        content += excel_note
        
        return content


class AutomatedEmailExecutionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AutomatedEmailExecution.objects.all()
    serializer_class = AutomatedEmailExecutionSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return AutomatedEmailExecution.objects.filter(schedule__created_by=self.request.user)


class CommandExecutionResultViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CommandExecutionResult.objects.all()
    serializer_class = CommandExecutionResultSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return CommandExecutionResult.objects.filter(execution__schedule__created_by=self.request.user)


    


class CommandTemplateViewSet(viewsets.ModelViewSet):
    queryset = CommandTemplate.objects.all()
    serializer_class = CommandTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return CommandTemplate.objects.filter(owner=self.request.user)

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)
