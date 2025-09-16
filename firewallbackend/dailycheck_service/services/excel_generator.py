"""
Générateur de fichiers Excel pour les rapports daily check
"""

import os
import base64
import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from django.conf import settings

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """
    Classe pour générer des fichiers Excel avec screenshots intégrés
    """
    
    def __init__(self):
        self.base_dir = os.path.expanduser('~/Documents/DailyCheck')
        os.makedirs(self.base_dir, exist_ok=True)
        self.workbook = None
        self.current_filepath = None
    
    def start_multi_firewall_report(self, firewalls: List, timestamp: str = None) -> str:
        """
        Démarre un nouveau rapport multi-firewall avec architecture de dossiers respectée
        
        Args:
            firewalls: Liste des firewalls pour déterminer la structure de dossiers
            timestamp: Timestamp pour le nom du fichier (optionnel)
            
        Returns:
            str: Chemin vers le fichier Excel
        """
        if timestamp is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Déterminer la structure de dossiers basée sur les firewalls
        # Si tous les firewalls sont du même datacenter et type, utiliser cette structure
        # Sinon, créer un dossier "Multi_DC" ou "Multi_Type"
        dc_names = set()
        fw_types = set()
        
        for firewall in firewalls:
            dc_name = firewall.data_center.name if firewall.data_center else 'Unknown_DC'
            fw_type = firewall.firewall_type.name if firewall.firewall_type else 'Unknown_FW_Type'
            dc_names.add(dc_name)
            fw_types.add(fw_type)
        
        # Déterminer le nom du dossier
        if len(dc_names) == 1 and len(fw_types) == 1:
            # Tous les firewalls sont du même datacenter et type
            dc_name = list(dc_names)[0]
            fw_type = list(fw_types)[0]
            folder_name = f"{dc_name}_{fw_type}"
        elif len(dc_names) == 1:
            # Même datacenter, types différents
            dc_name = list(dc_names)[0]
            folder_name = f"{dc_name}_Multi_Type"
        elif len(fw_types) == 1:
            # Même type, datacenters différents
            fw_type = list(fw_types)[0]
            folder_name = f"Multi_DC_{fw_type}"
        else:
            # Datacenters et types différents
            folder_name = "Multi_DC_Multi_Type"
        
        # Créer la structure de dossiers
        report_dir = os.path.join(self.base_dir, folder_name)
        os.makedirs(report_dir, exist_ok=True)
        
        # Créer le nom du fichier
        filename = f'daily_check_multi_{timestamp}.xlsx'
        self.current_filepath = os.path.join(report_dir, filename)
        
        # Créer le classeur Excel
        self.workbook = Workbook()
        ws = self.workbook.active
        ws.title = "Résumé"
        
        # Ajouter les informations générales sur la feuille de résumé
        self._add_summary_sheet(ws)
        
        return self.current_filepath
    
    def add_firewall_sheet(self, daily_check, check_results: List[Dict[str, Any]], firewall) -> None:
        """
        Ajoute une feuille pour un firewall au rapport existant
        
        Args:
            daily_check: Instance du modèle DailyCheck
            check_results: Résultats des commandes
            firewall: Instance du modèle Firewall
        """
        if not self.workbook:
            raise ValueError("Workbook not initialized. Call start_multi_firewall_report() first.")
        
        # Créer la feuille pour ce firewall
        sheet_name = self._get_sheet_name(firewall)
        ws = self.workbook.create_sheet(sheet_name)
        
        # Ajouter les informations du firewall
        self._add_firewall_info(ws, daily_check, firewall)
        
        # Ajouter le screenshot si disponible
        self._add_screenshot(ws, daily_check)
        
        # Ajouter les commandes et sorties
        self._add_commands_output(ws, check_results)
    
    def finalize_report(self) -> str:
        """
        Finalise et sauvegarde le rapport multi-firewall
        
        Returns:
            str: Chemin vers le fichier Excel généré
        """
        if not self.workbook or not self.current_filepath:
            raise ValueError("No report to finalize")
        
        # Sauvegarder le fichier
        self.workbook.save(self.current_filepath)
        logger.info(f"Multi-firewall Excel report generated: {self.current_filepath}")
        
        filepath = self.current_filepath
        # Réinitialiser pour le prochain rapport
        self.workbook = None
        self.current_filepath = None
        
        return filepath
    
    def _add_summary_sheet(self, ws):
        """
        Ajoute les informations générales sur la feuille de résumé
        
        Args:
            ws: Feuille de calcul de résumé
        """
        # Titre du rapport
        ws.cell(row=1, column=1, value="RAPPORT DAILY CHECK MULTI-FIREWALL").font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value=f"Date de génération: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(bold=True)
        
        # Informations sur la structure de dossiers
        folder_info = self.current_filepath.replace(self.base_dir, "").replace("\\", "/").strip("/")
        ws.cell(row=3, column=1, value=f"Emplacement: {folder_info}").font = Font(italic=True)
        
        # En-têtes du tableau de résumé
        headers = ["Firewall", "IP Address", "Type", "Data Center", "Statut", "Screenshot", "Feuille"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
    
    def update_summary_with_firewall(self, daily_check, firewall, row_number: int):
        """
        Met à jour la feuille de résumé avec les informations d'un firewall
        
        Args:
            daily_check: Instance du modèle DailyCheck
            firewall: Instance du modèle Firewall
            row_number: Numéro de ligne pour ajouter les informations
        """
        if not self.workbook:
            return
        
        ws = self.workbook["Résumé"]
        
        # Ajouter les informations du firewall
        ws.cell(row=row_number, column=1, value=firewall.name)
        ws.cell(row=row_number, column=2, value=firewall.ip_address)
        ws.cell(row=row_number, column=3, value=firewall.firewall_type.name if firewall.firewall_type else "Unknown")
        ws.cell(row=row_number, column=4, value=firewall.data_center.name if firewall.data_center else "Unknown")
        ws.cell(row=row_number, column=5, value=daily_check.status)
        ws.cell(row=row_number, column=6, value="✅" if daily_check.screenshot_captured else "❌")
        ws.cell(row=row_number, column=7, value=self._get_sheet_name(firewall))
    
    def generate_report(self, daily_check, check_results: List[Dict[str, Any]], firewall) -> str:
        """
        Génère un fichier Excel pour un daily check
        
        Args:
            daily_check: Instance du modèle DailyCheck
            check_results: Résultats des commandes
            firewall: Instance du modèle Firewall
            
        Returns:
            str: Chemin vers le fichier Excel généré
        """
        try:
            # Créer la structure de dossiers
            filepath = self._create_file_structure(firewall)
            
            # Créer le classeur Excel
            wb = Workbook()
            ws = wb.active
            
            # Supprimer la feuille par défaut
            wb.remove(ws)
            
            # Créer la feuille pour ce firewall
            sheet_name = self._get_sheet_name(firewall)
            ws = wb.create_sheet(sheet_name)
            
            # Ajouter les informations du firewall
            self._add_firewall_info(ws, daily_check, firewall)
            
            # Ajouter le screenshot si disponible
            self._add_screenshot(ws, daily_check)
            
            # Ajouter les commandes et sorties
            self._add_commands_output(ws, check_results)
            
            # Sauvegarder le fichier
            wb.save(filepath)
            logger.info(f"Excel report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise
    
    def _create_file_structure(self, firewall) -> str:
        """
        Crée la structure de dossiers et retourne le chemin du fichier
        
        Args:
            firewall: Instance du modèle Firewall
            
        Returns:
            str: Chemin vers le fichier Excel
        """
        # Créer les dossiers
        dc_name = firewall.data_center.name if firewall.data_center else 'Unknown_DC'
        fw_type = firewall.firewall_type.name if firewall.firewall_type else 'Unknown_FW_Type'
        
        dc_dir = os.path.join(self.base_dir, dc_name)
        fw_type_dir = os.path.join(dc_dir, fw_type)
        
        os.makedirs(dc_dir, exist_ok=True)
        os.makedirs(fw_type_dir, exist_ok=True)
        
        # Générer le nom du fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'daily_check_{timestamp}.xlsx'
        
        return os.path.join(fw_type_dir, filename)
    
    def _get_sheet_name(self, firewall) -> str:
        """
        Génère le nom de la feuille Excel
        
        Args:
            firewall: Instance du modèle Firewall
            
        Returns:
            str: Nom de la feuille (tronqué à 31 caractères)
        """
        sheet_name = f"{firewall.name}_{firewall.ip_address}"
        return sheet_name[:31]  # Limite Excel
    
    def _add_firewall_info(self, ws, daily_check, firewall):
        """
        Ajoute les informations du firewall en haut de la feuille
        
        Args:
            ws: Feuille de calcul
            daily_check: Instance du modèle DailyCheck
            firewall: Instance du modèle Firewall
        """
        info_data = [
            ("Firewall:", firewall.name),
            ("Adresse IP:", firewall.ip_address),
            ("Type:", firewall.firewall_type.name if firewall.firewall_type else "Unknown"),
            ("Data Center:", firewall.data_center.name if firewall.data_center else "Unknown"),
            ("Date de vérification:", daily_check.check_date.strftime('%Y-%m-%d %H:%M:%S')),
            ("Statut:", daily_check.status),
            ("Screenshot:", "✅ Capturé" if daily_check.screenshot_captured else "❌ Non capturé")
        ]
        
        # Écrire les informations à gauche
        for row, (label, value) in enumerate(info_data, 1):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
    
    def _add_screenshot(self, ws, daily_check):
        """
        Ajoute le screenshot en haut à droite de la feuille
        
        Args:
            ws: Feuille de calcul
            daily_check: Instance du modèle DailyCheck
        """
        if not daily_check.screenshot_captured or not daily_check.screenshot_base64:
            return
        
        try:
            # Décoder l'image base64
            image_data = base64.b64decode(daily_check.screenshot_base64)
            image = Image(io.BytesIO(image_data))
            
            # Redimensionner l'image (taille agrandie x2)
            max_width = 800
            max_height = 600
            if image.width > max_width or image.height > max_height:
                ratio = min(max_width / image.width, max_height / image.height)
                image.width = int(image.width * ratio)
                image.height = int(image.height * ratio)
            
            # Placer l'image en haut à droite (colonne E, ligne 1)
            ws.add_image(image, 'E1')
            
            # Ajuster la hauteur des lignes pour l'image agrandie
            for i in range(1, 12):  # Ajuster les 11 premières lignes pour l'image plus grande
                ws.row_dimensions[i].height = max(ws.row_dimensions[i].height or 15, int(image.height / 11))
            
            logger.info(f"Screenshot added to Excel")
            
        except Exception as e:
            logger.error(f"Error adding screenshot to Excel: {e}")
            ws.cell(row=7, column=1, value="⚠️ Erreur: Impossible d'ajouter le screenshot")
            ws.cell(row=7, column=1).font = Font(color="FF0000")
    
    def _add_commands_output(self, ws, check_results: List[Dict[str, Any]]):
        """
        Ajoute les commandes et sorties à la feuille
        
        Args:
            ws: Feuille de calcul
            check_results: Résultats des commandes
        """
        # Préparer les données
        all_output_data = []
        for cmd_result in check_results:
            all_output_data.append(f"COMMAND: {cmd_result['command']}")
            if cmd_result.get('output'):
                all_output_data.extend(cmd_result['output'].split('\n'))
            all_output_data.append("")
        
        # Commencer les commandes après les informations (ligne 10)
        start_row = 10
        for i, data in enumerate(all_output_data):
            ws.cell(row=start_row + i, column=1, value=data)
        
        # Formater les commandes (en rouge)
        red_font = Font(color="FF0000")
        for row in range(start_row, start_row + len(all_output_data)):
            cell = ws.cell(row=row, column=1)
            if isinstance(cell.value, str) and cell.value.startswith('COMMAND: '):
                cell.font = red_font
