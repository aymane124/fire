from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import DailyCheck, CheckCommand
from .serializers import DailyCheckSerializer, CheckCommandSerializer
from command_service.models import FirewallCommand
from command_service.views import FirewallCommandViewSet
import pandas as pd
from datetime import datetime
import os
import time
from django.http import HttpResponse
from django.conf import settings
import logging
import paramiko
from auth_service.models import SSHUser
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import threading
from queue import Queue
import json
import base64
import io
from .screenshot_integration import capture_firewall_screenshot_with_fallback

logger = logging.getLogger(__name__)

# Queue pour stocker les tâches en cours
task_queue = Queue()
# Dictionnaire pour stocker l'état des tâches
task_status = {}

def background_task_worker():
    while True:
        try:
            task = task_queue.get()
            if task is None:
                break
                
            task_id = task['task_id']
            task_status[task_id] = {
                'status': 'running',
                'progress': 0,
                'message': 'Starting daily checks...'
            }
            
            try:
                # Exécuter les daily checks
                firewalls = task['firewalls']
                commands = task['commands']
                user = task['user']
                
                # Créer le répertoire de base
                if os.path.exists('/app'):
                    documents_path = '/app/reports'
                else:
                    documents_path = os.path.expanduser('~/Documents')
                base_dir = os.path.join(documents_path, 'DailyCheck')
                os.makedirs(base_dir, exist_ok=True)
                
                # Grouper les firewalls par data center et type
                firewall_groups = {}
                for firewall in firewalls:
                    dc_name = firewall.data_center.name if firewall.data_center else 'Unknown_DC'
                    fw_type = firewall.firewall_type.name if firewall.firewall_type else 'Unknown_FW_Type'
                    
                    if dc_name not in firewall_groups:
                        firewall_groups[dc_name] = {}
                    if fw_type not in firewall_groups[dc_name]:
                        firewall_groups[dc_name][fw_type] = []
                    
                    firewall_groups[dc_name][fw_type].append(firewall)
                
                total_firewalls = len(firewalls)
                processed_firewalls = 0
                results = []
                
                for dc_name, fw_types in firewall_groups.items():
                    dc_dir = os.path.join(base_dir, dc_name)
                    os.makedirs(dc_dir, exist_ok=True)
                    
                    for fw_type, group_firewalls in fw_types.items():
                        fw_type_dir = os.path.join(dc_dir, fw_type)
                        os.makedirs(fw_type_dir, exist_ok=True)
                        
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        filename = f'daily_check_{timestamp}.xlsx'
                        filepath = os.path.join(fw_type_dir, filename)
                        
                        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                            for firewall in group_firewalls:
                                try:
                                    # Mettre à jour le statut
                                    task_status[task_id]['message'] = f'Processing {firewall.name}...'
                                    
                                    # Créer daily check
                                    daily_check = DailyCheck.objects.create(
                                        firewall=firewall,
                                        status='PENDING'
                                    )
                                    
                                    # Capturer le screenshot du dashboard
                                    logger.info(f"Capturing screenshot for firewall {firewall.name}")
                                    screenshot_result = capture_firewall_screenshot_with_fallback(firewall, user)
                                    
                                    if screenshot_result['success']:
                                        daily_check.screenshot_base64 = screenshot_result['screenshot_base64']
                                        daily_check.screenshot_captured = True
                                        logger.info(f"Screenshot captured for firewall {firewall.name}")
                                    else:
                                        logger.warning(f"Screenshot capture failed for firewall {firewall.name}: {screenshot_result.get('error', 'Unknown error')}")
                                    
                                    daily_check.save()
                                    
                                    # Exécuter les commandes
                                    check_results = []
                                    ssh_user = SSHUser.objects.get(user=user)
                                    decrypted_password = ssh_user.get_ssh_password()
                                    
                                    ssh = paramiko.SSHClient()
                                    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                                    ssh.connect(
                                        firewall.ip_address,
                                        username=ssh_user.ssh_username,
                                        password=decrypted_password,
                                        timeout=10
                                    )
                                    
                                    channel = ssh.invoke_shell()
                                    time.sleep(2)
                                    
                                    for cmd in commands:
                                        channel.send(cmd + '\n')
                                        time.sleep(2)
                                        
                                        output = ""
                                        start_time = time.time()
                                        while True:
                                            if channel.recv_ready():
                                                chunk = channel.recv(4096).decode('utf-8')
                                                output += chunk
                                                if '#' in chunk or '>' in chunk:
                                                    time.sleep(1)
                                                    if not channel.recv_ready():
                                                        break
                                                elif '--More--' in chunk:
                                                    channel.send(' ')
                                                    output = output.replace('--More--', '')
                                                    time.sleep(1)
                                            else:
                                                if time.time() - start_time > 30:
                                                    break
                                                time.sleep(0.1)
                                        
                                        # Nettoyer la sortie
                                        output_lines = output.split('\n')
                                        cleaned_output_lines = []
                                        for line in output_lines:
                                            if line.strip() != cmd.strip() and not line.strip().endswith(tuple(['#', '>'])):
                                                cleaned_output_lines.append(line)
                                        output = '\n'.join(cleaned_output_lines).strip()
                                        
                                        # Créer l'enregistrement
                                        command_result = CheckCommand.objects.create(
                                            daily_check=daily_check,
                                            command=cmd,
                                            actual_output=output,
                                            status='SUCCESS'
                                        )
                                        check_results.append(command_result)
                                    
                                    channel.close()
                                    ssh.close()
                                    
                                    # Créer la feuille Excel
                                    sheet_name = f"{firewall.name}_{firewall.ip_address}"
                                    sheet_name = sheet_name[:31]
                                    
                                    # Créer d'abord la feuille vide
                                    ws = writer.book.create_sheet(sheet_name)
                                    
                                    # Ajouter les informations du firewall en haut
                                    info_data = [
                                        ("Firewall:", firewall.name),
                                        ("Adresse IP:", firewall.ip_address),
                                        ("Type:", firewall.firewall_type.name if firewall.firewall_type else "Unknown"),
                                        ("Data Center:", firewall.data_center.name if firewall.data_center else "Unknown"),
                                        ("Date de vérification:", daily_check.check_date.strftime('%Y-%m-%d %H:%M:%S')),
                                        ("Statut:", daily_check.status),
                                        ("Screenshot:", "✅ Capturé" if daily_check.screenshot_captured else "❌ Non capturé")
                                    ]
                                    
                                    # Écrire les informations à gauche
                                    for row, (label, value) in enumerate(info_data, 1):
                                        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                                        ws.cell(row=row, column=2, value=value)
                                    
                                    # Ajouter le screenshot en haut à droite si disponible
                                    if daily_check.screenshot_captured and daily_check.screenshot_base64:
                                        try:
                                            # Décoder l'image base64
                                            image_data = base64.b64decode(daily_check.screenshot_base64)
                                            image = Image(io.BytesIO(image_data))
                                            
                                            # Redimensionner l'image pour qu'elle rentre bien
                                            max_width = 400
                                            max_height = 300
                                            if image.width > max_width or image.height > max_height:
                                                ratio = min(max_width / image.width, max_height / image.height)
                                                image.width = int(image.width * ratio)
                                                image.height = int(image.height * ratio)
                                            
                                            # Placer l'image en haut à droite (colonne E, ligne 1)
                                            ws.add_image(image, 'E1')
                                            
                                            # Ajuster la hauteur des lignes pour l'image
                                            for i in range(1, 8):  # Ajuster les 7 premières lignes
                                                ws.row_dimensions[i].height = max(ws.row_dimensions[i].height or 15, int(image.height / 7))
                                            
                                            logger.info(f"Screenshot added to top-right of Excel for firewall {firewall.name}")
                                            
                                        except Exception as e:
                                            logger.error(f"Error adding screenshot to Excel for firewall {firewall.name}: {e}")
                                            ws.cell(row=7, column=1, value="⚠️ Erreur: Impossible d'ajouter le screenshot")
                                            ws.cell(row=7, column=1).font = Font(color="FF0000")
                                    
                                    # Ajouter les commandes et sorties en dessous
                                    all_output_data = []
                                    for cmd_result in check_results:
                                        all_output_data.append(f"COMMAND: {cmd_result.command}")
                                        if cmd_result.actual_output:
                                            all_output_data.extend(cmd_result.actual_output.split('\n'))
                                        all_output_data.append("")
                                    
                                    # Commencer les commandes après les informations (ligne 10)
                                    start_row = 10
                                    for i, data in enumerate(all_output_data):
                                        ws.cell(row=start_row + i, column=1, value=data)
                                    
                                    # Formater les commandes (en rouge)
                                    red_font = Font(color="FF0000")
                                    for row in range(start_row, start_row + len(all_output_data)):
                                        cell = ws.cell(row=row, column=1)
                                        if isinstance(cell.value, str) and cell.value.startswith('COMMAND: '):
                                            cell.font = red_font
                                    
                                    # Mettre à jour le statut
                                    daily_check.excel_report = filepath
                                    daily_check.status = 'SUCCESS'
                                    daily_check.save()
                                    
                                    results.append({
                                        'firewall_id': firewall.id,
                                        'status': 'SUCCESS',
                                        'success': True,
                                        'report_path': filepath
                                    })
                                    
                                except Exception as e:
                                    logger.error(f"Error processing firewall {firewall.id}: {str(e)}")
                                    results.append({
                                        'firewall_id': firewall.id,
                                        'status': 'FAILED',
                                        'success': False,
                                        'error': str(e)
                                    })
                                
                                processed_firewalls += 1
                                task_status[task_id]['progress'] = int((processed_firewalls / total_firewalls) * 100)
                
                task_status[task_id].update({
                    'status': 'completed',
                    'progress': 100,
                    'message': 'All daily checks completed',
                    'results': results
                })
                
            except Exception as e:
                logger.error(f"Error in background task: {str(e)}")
                task_status[task_id].update({
                    'status': 'failed',
                    'message': str(e)
                })
            
            finally:
                task_queue.task_done()
                
        except Exception as e:
            logger.error(f"Error in worker thread: {str(e)}")
            continue

# Démarrer le worker thread
worker_thread = threading.Thread(target=background_task_worker, daemon=True)
worker_thread.start()

class DailyCheckViewSet(viewsets.ModelViewSet):
    serializer_class = DailyCheckSerializer

    def get_queryset(self):
        # Get all daily checks and filter in Python
        daily_checks = DailyCheck.objects.all()
        user_checks = []
        
        for check in daily_checks:
            # Check if any history entry contains the current user
            if check.historique_dailycheck and 'entries' in check.historique_dailycheck:
                for entry in check.historique_dailycheck['entries']:
                    if entry.get('user') == str(self.request.user):
                        user_checks.append(check)
                        break
        
        return DailyCheck.objects.filter(id__in=[check.id for check in user_checks])

    def perform_create(self, serializer):
        instance = serializer.save()
        instance.add_to_history(
            action='create',
            status='success',
            details=f"Daily check created for firewall {instance.firewall.name}",
            user=self.request.user,
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        instance.add_to_history(
            action='update',
            status='success',
            details=f"Daily check updated for firewall {instance.firewall.name}",
            user=self.request.user,
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_destroy(self, instance):
        instance.add_to_history(
            action='delete',
            status='success',
            details=f"Daily check deleted for firewall {instance.firewall.name}",
            user=self.request.user,
            ip_address=self.request.META.get('REMOTE_ADDR')
        )
        instance.delete()

    def execute_commands_in_session(self, firewall, commands, user):
        """
        Exécute une série de commandes dans une seule session SSH
        """
        try:
            # Récupérer l'utilisateur SSH
            ssh_user = SSHUser.objects.get(user=user)
            decrypted_password = ssh_user.get_ssh_password()

            # Établir la connexion SSH
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                firewall.ip_address,
                username=ssh_user.ssh_username,
                password=decrypted_password,
                timeout=10
            )

            results = []
            # Créer un canal interactif
            channel = ssh.invoke_shell()
            time.sleep(2)  # Attendre que le shell soit prêt

            # Fonction pour lire la sortie complète
            def read_output(channel, timeout=30):
                output = ""
                start_time = time.time()
                
                while True:
                    # Vérifier si le canal est prêt à être lu
                    if channel.recv_ready():
                        chunk = channel.recv(4096).decode('utf-8')
                        output += chunk
                        
                        # Détecter le prompt final ou --More--
                        if '#' in chunk or '>' in chunk:
                             time.sleep(1)  # Attendre un peu plus pour être sûr
                             if not channel.recv_ready():
                                break
                        elif '--More--' in chunk:
                             # Envoyer un espace pour paginer
                             channel.send(' ')
                             # Supprimer --More-- de la sortie capturée pour la netteté
                             output = output.replace('--More--', '')
                             time.sleep(1) # Donner au pare-feu le temps de répondre

                    else:
                        # Vérifier le timeout
                        if time.time() - start_time > timeout:
                            logger.warning(f"Timeout reached while reading command output")
                            break
                        time.sleep(0.1)
                
                return output

            # Exécuter chaque commande
            for cmd in commands:
                try:
                    logger.info(f"Executing command: {cmd}")
                    # Envoyer la commande
                    channel.send(cmd + '\n')
                    time.sleep(2)  # Attendre que la commande commence

                    # Lire la sortie complète
                    output = read_output(channel)
                    logger.info(f"Command output length: {len(output)}")

                    # Nettoyer la sortie
                    # Supprimer la commande elle-même de la sortie et les prompts résiduels
                    output_lines = output.split('\n')
                    cleaned_output_lines = []
                    for line in output_lines:
                         # Ignorer les lignes qui contiennent uniquement la commande ou le prompt
                         if line.strip() != cmd.strip() and not line.strip().endswith(tuple(['#', '>'])):
                              cleaned_output_lines.append(line)
                    output = '\n'.join(cleaned_output_lines).strip()

                    # Créer l'enregistrement FirewallCommand
                    command_obj = FirewallCommand.objects.create(
                        firewall=firewall,
                        user=user,
                        command=cmd,
                        status='completed',
                        output=output
                    )

                    results.append({
                        'command': cmd,
                        'status': 'completed',
                        'output': output,
                        'error': None
                    })

                except Exception as e:
                    logger.error(f"Error executing command {cmd}: {str(e)}")
                    results.append({
                        'command': cmd,
                        'status': 'failed',
                        'output': None,
                        'error': str(e)
                    })

            # Fermer la connexion
            channel.close()
            ssh.close()

            return results

        except Exception as e:
            logger.error(f"SSH connection error: {str(e)}")
            raise

    @action(detail=True, methods=['post'])
    def run_commands(self, request, pk=None):
        daily_check = self.get_object()
        firewall = daily_check.firewall
        
        try:
            # Get commands from request body
            commands = request.data.get('commands', [])
            
            # Validate commands
            if not commands:
                return Response(
                    {'error': 'No commands provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not isinstance(commands, list):
                return Response(
                    {'error': 'Commands must be provided as a list'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Capturer le screenshot du dashboard avant d'exécuter les commandes
            logger.info(f"Capturing screenshot for firewall {firewall.name}")
            screenshot_result = capture_firewall_screenshot_with_fallback(firewall, request.user)
            
            if screenshot_result['success']:
                daily_check.screenshot_base64 = screenshot_result['screenshot_base64']
                daily_check.screenshot_captured = True
                logger.info(f"Screenshot captured for firewall {firewall.name}")
            else:
                logger.warning(f"Screenshot capture failed for firewall {firewall.name}: {screenshot_result.get('error', 'Unknown error')}")
            
            daily_check.save()
            
            # Execute commands in a single session
            results = self.execute_commands_in_session(firewall, commands, request.user)
            
            # Create CheckCommand records
            check_results = []
            for cmd_result in results:
                command_result = CheckCommand.objects.create(
                    daily_check=daily_check,
                    command=cmd_result['command'],
                    actual_output=cmd_result['output'] if cmd_result['status'] == 'completed' else cmd_result['error'],
                    status='SUCCESS' if cmd_result['status'] == 'completed' else 'FAILED'
                )
                command_result.add_to_history(
                    action='create',
                    status='success' if cmd_result['status'] == 'completed' else 'failed',
                    details=f"Command executed: {cmd_result['command']}",
                    user=request.user,
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                check_results.append(command_result)
            
            # Create Excel report
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'daily_check_{timestamp}.xlsx'
            
            # Create directories
            if os.path.exists('/app'):
                documents_path = '/app/reports'
            else:
                documents_path = os.path.expanduser('~/Documents')
            base_dir = os.path.join(documents_path, 'DailyCheck')
            os.makedirs(base_dir, exist_ok=True)
            
            if not hasattr(firewall, 'data_center') or not firewall.data_center:
                firewall.refresh_from_db()
            if not hasattr(firewall, 'firewall_type') or not firewall.firewall_type:
                firewall.refresh_from_db()

            dc_dir = os.path.join(base_dir, firewall.data_center.name if firewall.data_center else 'Unknown_DC')
            os.makedirs(dc_dir, exist_ok=True)
            
            fw_type_dir = os.path.join(dc_dir, firewall.firewall_type.name if firewall.firewall_type else 'Unknown_FW_Type')
            os.makedirs(fw_type_dir, exist_ok=True)
            
            filepath = os.path.join(fw_type_dir, filename)
            
            # Create Excel report with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create a sheet for this firewall
                sheet_name = f"{firewall.name}_{firewall.ip_address}"
                # Truncate sheet name if too long (Excel has a 31 character limit)
                sheet_name = sheet_name[:31]
                
                # Créer d'abord la feuille vide
                ws = writer.book.create_sheet(sheet_name)
                
                # Ajouter les informations du firewall en haut
                info_data = [
                    ("Firewall:", firewall.name),
                    ("Adresse IP:", firewall.ip_address),
                    ("Type:", firewall.firewall_type.name if firewall.firewall_type else "Unknown"),
                    ("Data Center:", firewall.data_center.name if firewall.data_center else "Unknown"),
                    ("Date de vérification:", daily_check.check_date.strftime('%Y-%m-%d %H:%M:%S')),
                    ("Statut:", daily_check.status),
                    ("Screenshot:", "✅ Capturé" if daily_check.screenshot_captured else "❌ Non capturé")
                ]
                
                # Écrire les informations à gauche
                for row, (label, value) in enumerate(info_data, 1):
                    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=value)
                
                # Ajouter le screenshot en haut à droite si disponible
                if daily_check.screenshot_captured and daily_check.screenshot_base64:
                    try:
                        # Décoder l'image base64
                        image_data = base64.b64decode(daily_check.screenshot_base64)
                        image = Image(io.BytesIO(image_data))
                        
                        # Redimensionner l'image pour qu'elle rentre bien
                        max_width = 400
                        max_height = 300
                        if image.width > max_width or image.height > max_height:
                            ratio = min(max_width / image.width, max_height / image.height)
                            image.width = int(image.width * ratio)
                            image.height = int(image.height * ratio)
                        
                        # Placer l'image en haut à droite (colonne E, ligne 1)
                        ws.add_image(image, 'E1')
                        
                        # Ajuster la hauteur des lignes pour l'image
                        for i in range(1, 8):  # Ajuster les 7 premières lignes
                            ws.row_dimensions[i].height = max(ws.row_dimensions[i].height or 15, int(image.height / 7))
                        
                        logger.info(f"Screenshot added to top-right of Excel for firewall {firewall.name}")
                        
                    except Exception as e:
                        logger.error(f"Error adding screenshot to Excel for firewall {firewall.name}: {e}")
                        ws.cell(row=7, column=1, value="⚠️ Erreur: Impossible d'ajouter le screenshot")
                        ws.cell(row=7, column=1).font = Font(color="FF0000")
                
                # Préparer les données des commandes
                all_output_data = []
                for cmd_result in check_results:
                    # Add command line with prefix
                    all_output_data.append(f"COMMAND: {cmd_result.command}")
                    # Add output
                    if cmd_result.actual_output:
                        all_output_data.extend(cmd_result.actual_output.split('\n'))
                    # Add empty line to separate commands
                    all_output_data.append("")
                
                # Commencer les commandes après les informations (ligne 10)
                start_row = 10
                for i, data in enumerate(all_output_data):
                    ws.cell(row=start_row + i, column=1, value=data)

                # Formater les commandes (en rouge)
                red_font = Font(color="FF0000")
                for row in range(start_row, start_row + len(all_output_data)):
                    cell = ws.cell(row=row, column=1)
                    if isinstance(cell.value, str) and cell.value.startswith('COMMAND: '):
                        cell.font = red_font

            # Update daily check status
            daily_check.excel_report = filepath
            daily_check.status = 'SUCCESS'
            for res in check_results:
                if res.status == 'FAILED':
                    daily_check.status = 'FAILED'
                    break

            daily_check.save()
            daily_check.add_to_history(
                action='run_commands',
                status='success' if daily_check.status == 'SUCCESS' else 'failed',
                details=f"Daily check completed with {len(check_results)} commands",
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return Response({
                'status': daily_check.status,
                'message': 'Daily check completed and report generated',
                'report_path': filepath
            })
            
        except Exception as e:
            logger.error(f"Error in run_commands: {str(e)}")
            daily_check.status = 'FAILED'
            daily_check.notes = str(e)
            daily_check.save()
            daily_check.add_to_history(
                action='run_commands',
                status='failed',
                details=f"Error during daily check: {str(e)}",
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def get_reports(self, request):
        reports = DailyCheck.objects.filter(excel_report__isnull=False)
        serializer = self.get_serializer(reports, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def download_report(self, request, pk=None):
        daily_check = self.get_object()
        if not daily_check.excel_report:
            return Response({
                'error': 'No report available for this daily check'
            }, status=status.HTTP_404_NOT_FOUND)

        try:
            with open(daily_check.excel_report, 'rb') as file:
                file_content = file.read()
            response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(daily_check.excel_report)}"'
            
            daily_check.add_to_history(
                action='download_report',
                status='success',
                details=f"Report downloaded: {os.path.basename(daily_check.excel_report)}",
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return response
        except Exception as e:
            daily_check.add_to_history(
                action='download_report',
                status='failed',
                details=f"Error downloading report: {str(e)}",
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def run_multiple_checks(self, request):
        try:
            firewall_ids = request.data.get('firewalls', [])
            commands = request.data.get('commands', [])
            
            if not firewall_ids or not commands:
                return Response(
                    {'error': 'No firewalls or commands provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Créer un ID unique pour la tâche
            task_id = f"task_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Récupérer les firewalls
            from command_service.models import Firewall
            firewalls = list(Firewall.objects.filter(id__in=firewall_ids))
            
            # Ajouter la tâche à la queue
            task_queue.put({
                'task_id': task_id,
                'firewalls': firewalls,
                'commands': commands,
                'user': request.user
            })
            
            # Initialiser le statut de la tâche
            task_status[task_id] = {
                'status': 'pending',
                'progress': 0,
                'message': 'Task queued'
            }
            
            return Response({
                'status': 'success',
                'message': 'Daily checks started',
                'task_id': task_id
            })
            
        except Exception as e:
            logger.error(f"Error in run_multiple_checks: {str(e)}")
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def check_task_status(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({
                'error': 'No task ID provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if task_id not in task_status:
            return Response({
                'error': 'Task not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        return Response(task_status[task_id])

    @action(detail=False, methods=['get'])
    def download_multiple_reports(self, request):
        try:
            # Get the report path from query parameters
            report_path = request.query_params.get('report_path')
            if not report_path:
                return Response({
                    'error': 'No report path provided'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if file exists
            if not os.path.exists(report_path):
                return Response({
                    'error': 'Report file not found'
                }, status=status.HTTP_404_NOT_FOUND)

            # Create file response
            with open(report_path, 'rb') as file:
                file_content = file.read()
            response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(report_path)}"'
            
            # Add to history for all related daily checks
            daily_checks = DailyCheck.objects.filter(excel_report=report_path)
            for daily_check in daily_checks:
                daily_check.add_to_history(
                    action='download_multiple_reports',
                    status='success',
                    details=f"Multiple reports downloaded: {os.path.basename(report_path)}",
                    user=request.user,
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            
            return response
        except Exception as e:
            logger.error(f"Error in download_multiple_reports: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 